#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Export item_tbitemconfig from local AssetBundle (same table GM store uses).

GM store builds its list from ConfigManager.GetTbItemConfig().DataList, loaded from
Assets/Res/Config/ExcelGeneral/item_tbitemconfig.bytes inside bundle
4bd60e623f3f8796cb234b3f01f0c91a.b.

当前客户端为 Luban Bright.Serialization.ByteBuf 二进制（表头 ReadSize 条目数，
再按 ItemConfig 字段顺序反序列化）。旧版 aa38+定长 124B 布局已废弃。
"""
from __future__ import annotations

import argparse
import csv
import json
import struct
import sys
from datetime import datetime
from pathlib import Path

import re
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
DEFAULT_BUNDLE = ROOT / "cg37_Data/assets/4bd60e623f3f8796cb234b3f01f0c91a.b"
DEFAULT_OUT = ROOT.parent / "发布plugin" / "道具配表_Id名称对照.xlsx"
_LEVEL_ONLY = re.compile(r"^\d+级$")
_HAS_CJK = re.compile(r"[\u4e00-\u9fff]")

# 与 hotfix ItemConfig(ByteBuf) 一致：三字符串后 BatchUse → Id → … → Resource 列表
_INT_FIELDS_AFTER_LABEL = [
    "BatchUse",
    "Id",
    "Imagenumber",
    "Cost",
    "Type",
    "Quality",
    "Rank",
    "Bothhand",
    "Fieldtype",
    "Ableusebattle",
    "Target",
    "Maxremain",
    "Level",
    "Basefailedprob",
    "DurabilityMin",
    "DurabilityMax",
    "AttacknumMin",
    "AttacknumMax",
    "Ableeffectbetweenhave",
    "Modflg",
    "AttackMin",
    "AttackMax",
    "DefenceMin",
    "DefenceMax",
    "AgilityMin",
    "AgilityMax",
    "MagicMin",
    "MagicMax",
    "RecoveryMin",
    "RecoveryMax",
    "CriticalMin",
    "CriticalMax",
    "CounterMin",
    "CounterMax",
    "HitrateMin",
    "HitrateMax",
    "AvoidMin",
    "AvoidMax",
    "HpMin",
    "HpMax",
    "FpMin",
    "FpMax",
    "LuckMin",
    "LuckMax",
    "CharismaMin",
    "CharismaMax",
    "CharmMin",
    "CharmMax",
    "Attrib",
    "Attrib2",
    "Attribvalue",
    "Attribvalue2",
    "StaminaMin",
    "StaminaMax",
    "DexMin",
    "DexMax",
    "IntelligenceMin",
    "IntelligenceMax",
    "PoisonMin",
    "PoisonMax",
    "SleepMin",
    "SleepMax",
    "StoneMin",
    "StoneMax",
    "DrunkMin",
    "DrunkMax",
    "ConfusionMin",
    "ConfusionMax",
    "AmnesiaMin",
    "AmnesiaMax",
    "Specialeffect",
    "Specialeffectvalue",
    "Specialeffectvalue2",
    "MatWeapon",
    "MatArmour",
    "MatAccessory",
    "Useaction",
]


class ByteBuf:
    """Bright.Serialization.ByteBuf 精简读实现（ReadUint / ReadString / ReadBool）。"""

    __slots__ = ("_data", "_i")

    def __init__(self, data: bytes, offset: int = 0) -> None:
        self._data = data
        self._i = offset

    @property
    def position(self) -> int:
        return self._i

    @property
    def remaining(self) -> int:
        return len(self._data) - self._i

    def read_uint(self) -> int:
        data = self._data
        i = self._i
        num = data[i]
        if num < 128:
            self._i = i + 1
            return num
        if num < 192:
            self._i = i + 2
            return ((num & 0x3F) << 8) | data[i + 1]
        if num < 224:
            self._i = i + 3
            return ((num & 0x1F) << 16) | (data[i + 1] << 8) | data[i + 2]
        if num < 240:
            self._i = i + 4
            return (
                ((num & 0xF) << 24)
                | (data[i + 1] << 16)
                | (data[i + 2] << 8)
                | data[i + 3]
            )
        self._i = i + 5
        return (
            (data[i + 1] << 24)
            | (data[i + 2] << 16)
            | (data[i + 3] << 8)
            | data[i + 4]
        ) & 0xFFFFFFFF

    def read_int(self) -> int:
        u = self.read_uint()
        return u - 0x100000000 if u >= 0x80000000 else u

    def read_size(self) -> int:
        return self.read_uint()

    def read_string(self) -> str:
        n = self.read_size()
        if n <= 0:
            return ""
        end = self._i + n
        if end > len(self._data):
            raise ValueError(f"string overruns buffer at {self._i}, n={n}")
        s = self._data[self._i : end].decode("utf-8")
        self._i = end
        return s

    def read_bool(self) -> bool:
        v = self._data[self._i]
        self._i += 1
        return v != 0


def extract_textasset_bytes(bundle_path: Path, asset_name: str) -> bytes:
    try:
        import UnityPy
    except ImportError as exc:
        raise SystemExit("UnityPy required: pip install UnityPy") from exc

    raw = bundle_path.read_bytes()
    idx = raw.find(b"UnityFS")
    if idx < 0:
        raise FileNotFoundError(f"UnityFS header not found in {bundle_path}")
    env = UnityPy.load(raw[idx:])
    for obj in env.objects:
        if obj.type.name != "TextAsset":
            continue
        data = obj.read()
        if data.m_Name != asset_name:
            continue
        reader = obj.reader
        reader.Position = obj.byte_start
        rawobj = reader.read(obj.byte_size)
        pos = 4 + struct.unpack_from("<I", rawobj, 0)[0]
        pos = (pos + 3) & ~3
        size = struct.unpack_from("<I", rawobj, pos)[0]
        pos += 4
        return rawobj[pos : pos + size]
    raise FileNotFoundError(f"{asset_name} not in {bundle_path}")


def _clean_text(value: str) -> str:
    value = value.encode("utf-8", "replace").decode("utf-8")
    return "".join(ch for ch in value if ch >= " " or ch in "\t")


def _valid_secret(name: str) -> bool:
    if not name or len(name) > 40:
        return False
    if _LEVEL_ONLY.match(name):
        return False
    if name.endswith("？") or name.endswith("?"):
        return False
    return bool(_HAS_CJK.search(name))


def _read_item_config(buf: ByteBuf) -> dict:
    name = buf.read_string()
    secret = buf.read_string()
    label = buf.read_string()
    fields: dict[str, object] = {
        "Name": _clean_text(name),
        "Secretname": _clean_text(secret),
        "Label": _clean_text(label),
    }
    for key in _INT_FIELDS_AFTER_LABEL:
        fields[key] = buf.read_int()
    fields["Dropatlogout"] = buf.read_bool()
    fields["Vanishatdrop"] = buf.read_bool()
    fields["Canpetmail"] = buf.read_bool()
    fields["RssMin"] = buf.read_int()
    fields["RssMax"] = buf.read_int()
    fields["Cansell"] = buf.read_bool()
    for key in ("Exp1", "Exp2", "Rareflg", "Inboxflg", "AdmMin", "AdmMax", "Sellunit"):
        fields[key] = buf.read_int()
    n = buf.read_size()
    if n < 0 or n > 10_000:
        raise ValueError(f"Resource list too large: {n}")
    fields["Resource"] = [buf.read_int() for _ in range(n)]
    return fields


# 导出用字段（中文列名, 宽度）— 基础 + 常用详细属性
EXPORT_FIELDS: list[tuple[str, str, int]] = [
    ("Id", "道具ID", 12),
    ("Secretname", "显示名", 28),
    ("Name", "分类/内部名", 16),
    ("Label", "标签", 10),
    ("Type", "类型Type", 10),
    ("Level", "等级", 8),
    ("Quality", "品质", 8),
    ("Rank", "Rank", 8),
    ("Cost", "价格Cost", 10),
    ("Imagenumber", "图标编号", 10),
    ("BatchUse", "批量使用", 10),
    ("Fieldtype", "使用场景", 10),
    ("Ableusebattle", "战斗可用", 10),
    ("Target", "目标", 8),
    ("Maxremain", "最大堆叠", 10),
    ("Bothhand", "双手", 8),
    ("DurabilityMin", "耐久Min", 10),
    ("DurabilityMax", "耐久Max", 10),
    ("AttackMin", "攻击Min", 10),
    ("AttackMax", "攻击Max", 10),
    ("DefenceMin", "防御Min", 10),
    ("DefenceMax", "防御Max", 10),
    ("AgilityMin", "敏捷Min", 10),
    ("AgilityMax", "敏捷Max", 10),
    ("MagicMin", "精神Min", 10),
    ("MagicMax", "精神Max", 10),
    ("RecoveryMin", "回复Min", 10),
    ("RecoveryMax", "回复Max", 10),
    ("CriticalMin", "必杀Min", 10),
    ("CriticalMax", "必杀Max", 10),
    ("CounterMin", "反击Min", 10),
    ("CounterMax", "反击Max", 10),
    ("HitrateMin", "命中Min", 10),
    ("HitrateMax", "命中Max", 10),
    ("AvoidMin", "闪躲Min", 10),
    ("AvoidMax", "闪躲Max", 10),
    ("HpMin", "HP Min", 10),
    ("HpMax", "HP Max", 10),
    ("FpMin", "FP Min", 10),
    ("FpMax", "FP Max", 10),
    ("LuckMin", "幸运Min", 10),
    ("LuckMax", "幸运Max", 10),
    ("CharmMin", "魅力Min", 10),
    ("CharmMax", "魅力Max", 10),
    ("StaminaMin", "体力Min", 10),
    ("StaminaMax", "体力Max", 10),
    ("DexMin", "腕力/Dex Min", 12),
    ("DexMax", "腕力/Dex Max", 12),
    ("IntelligenceMin", "智力Min", 10),
    ("IntelligenceMax", "智力Max", 10),
    ("Attrib", "属性1", 8),
    ("Attribvalue", "属性1值", 10),
    ("Attrib2", "属性2", 8),
    ("Attribvalue2", "属性2值", 10),
    ("PoisonMin", "毒Min", 8),
    ("PoisonMax", "毒Max", 8),
    ("SleepMin", "睡Min", 8),
    ("SleepMax", "睡Max", 8),
    ("StoneMin", "石Min", 8),
    ("StoneMax", "石Max", 8),
    ("DrunkMin", "醉Min", 8),
    ("DrunkMax", "醉Max", 8),
    ("ConfusionMin", "乱Min", 8),
    ("ConfusionMax", "乱Max", 8),
    ("AmnesiaMin", "忘Min", 8),
    ("AmnesiaMax", "忘Max", 8),
    ("Specialeffect", "特效ID", 10),
    ("Specialeffectvalue", "特效值1", 10),
    ("Specialeffectvalue2", "特效值2", 10),
    ("Useaction", "使用动作", 10),
    ("Cansell", "可出售", 8),
    ("Dropatlogout", "下线掉落", 10),
    ("Vanishatdrop", "丢弃消失", 10),
    ("Canpetmail", "可邮寄", 8),
    ("Rareflg", "稀有标记", 10),
    ("Sellunit", "出售单位", 10),
    ("Modflg", "Modflg", 10),
]

EXCEL_COLUMNS = list(EXPORT_FIELDS)

NAME_INDEX_COLUMNS = [
    ("Secretname", "显示名", 28),
    ("Id", "道具ID", 12),
    ("Name", "分类/内部名", 16),
    ("Label", "标签", 10),
    ("Type", "类型Type", 10),
    ("Level", "等级", 8),
    ("AttackMin", "攻击Min", 10),
    ("AttackMax", "攻击Max", 10),
    ("DefenceMin", "防御Min", 10),
    ("DefenceMax", "防御Max", 10),
    ("Specialeffect", "特效ID", 10),
]

_KEEP_KEYS = {key for key, _, _ in EXCEL_COLUMNS}


def load_items(config_bytes: bytes) -> list[dict]:
    """按 Luban TbItemConfig 格式解析；返回去重后的导出行（含详细属性）。"""
    buf = ByteBuf(config_bytes)
    count = buf.read_size()
    if count <= 0 or count > 500_000:
        raise ValueError(f"suspicious item count: {count}")

    items_by_id: dict[int, dict] = {}
    for idx in range(count):
        try:
            row = _read_item_config(buf)
        except Exception as exc:
            raise ValueError(f"parse failed at item #{idx}, pos={buf.position}: {exc}") from exc
        item_id = int(row["Id"])  # type: ignore[arg-type]
        secret = str(row["Secretname"])
        if item_id <= 0 or not _valid_secret(secret):
            continue
        export_row = {k: row[k] for k in _KEEP_KEYS if k in row}
        for bk in ("Cansell", "Dropatlogout", "Vanishatdrop", "Canpetmail"):
            if bk in export_row:
                export_row[bk] = "是" if export_row[bk] else "否"
        items_by_id.setdefault(item_id, export_row)

    items = list(items_by_id.values())
    items.sort(key=lambda x: int(x["Id"]))
    return items


def _style_header_row(ws, row: int, columns: list[tuple[str, str, int]]) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(name="微软雅黑", bold=True, color="FFFFFF", size=11)
    thin = Side(style="thin", color="D0D7DE")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")
    for col_idx, (_, title_cn, width) in enumerate(columns, start=1):
        cell = ws.cell(row=row, column=col_idx, value=title_cn)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = border
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[row].height = 22


def _fill_item_rows(
    ws,
    items: list[dict],
    columns: list[tuple[str, str, int]],
    *,
    header_row: int,
) -> None:
    body_font = Font(name="微软雅黑", size=10)
    alt_fill = PatternFill("solid", fgColor="F2F7FB")
    white_fill = PatternFill("solid", fgColor="FFFFFF")
    thin = Side(style="thin", color="D0D7DE")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")

    for row_idx, item in enumerate(items, start=header_row + 1):
        fill = alt_fill if (row_idx - header_row) % 2 == 0 else white_fill
        for col_idx, (key, _, _) in enumerate(columns, start=1):
            value = item.get(key, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = body_font
            cell.fill = fill
            cell.border = border
            if key in ("Id", "Imagenumber", "BatchUse") or (
                isinstance(value, int) and key not in ("Secretname", "Name", "Label")
            ):
                cell.alignment = center
                if isinstance(value, int):
                    cell.number_format = "0"
            else:
                cell.alignment = left

    last = header_row + len(items)
    ws.auto_filter.ref = f"A{header_row}:{get_column_letter(len(columns))}{last}"


def _write_readme_sheet(wb: Workbook, item_count: int) -> None:
    ws = wb.create_sheet("说明", 0)
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 72

    title = ws.cell(row=1, column=1, value="道具配表说明")
    title.font = Font(name="微软雅黑", bold=True, size=16, color="1F4E79")
    ws.merge_cells("A1:B1")
    ws.row_dimensions[1].height = 30

    rows = [
        ("用途", "Id ↔ 显示名 + 配表详细属性。背包实例数值（当前耐久等）仍以服务端同步为准。"),
        ("显示名", "配表 Secretname；一般等于 UI / 背包看到的中文名。"),
        ("详细属性", "见「详细属性」表：攻防、HP/FP、异常抗性、特效、可否出售等，均来自本地 ItemConfig。"),
        ("游戏内查看", "背包点道具 → 详情/Tip；客户端用 ItemConfig + 格子实例数据拼出来。"),
        ("代码查看", "Manager<ConfigManager>.Instance.GetTbItemConfig().GetOrDefault(itemId)"),
        ("查法", "「按名称」搜中文；「详细属性」看完整列；可用筛选/冻结首行。"),
        ("导出时间", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ("条目数", str(item_count)),
    ]
    label_font = Font(name="微软雅黑", bold=True, size=10, color="1F4E79")
    body_font = Font(name="微软雅黑", size=10)
    label_fill = PatternFill("solid", fgColor="E8F1F8")
    thin = Side(style="thin", color="D0D7DE")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for i, (k, v) in enumerate(rows, start=3):
        c1 = ws.cell(row=i, column=1, value=k)
        c2 = ws.cell(row=i, column=2, value=v)
        c1.font = label_font
        c1.fill = label_fill
        c1.border = border
        c1.alignment = Alignment(horizontal="center", vertical="center")
        c2.font = body_font
        c2.border = border
        c2.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws.row_dimensions[i].height = 36 if i <= 8 else 22


def write_excel(items: list[dict], path: Path) -> None:
    wb = Workbook()
    # remove default; rebuild ordered sheets
    default = wb.active
    wb.remove(default)

    _write_readme_sheet(wb, len(items))

    left = Alignment(horizontal="left", vertical="center")

    # Sheet: full details (same as 按ID but clearer title)
    ws_id = wb.create_sheet("详细属性")
    ws_id.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(EXCEL_COLUMNS))
    title = ws_id.cell(row=1, column=1, value="魔力宝贝：序章 — 道具详细属性（本地 ItemConfig，按道具ID排序）")
    title.font = Font(name="微软雅黑", bold=True, size=14, color="1F4E79")
    title.alignment = left
    ws_id.row_dimensions[1].height = 28
    meta = ws_id.cell(
        row=2,
        column=1,
        value=f"导出时间：{datetime.now():%Y-%m-%d %H:%M:%S}    条目数：{len(items)}    来源：item_tbitemconfig（攻击/防御等为配表区间，实例值可能在区间内）",
    )
    ws_id.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(EXCEL_COLUMNS))
    meta.font = Font(name="微软雅黑", size=9, color="666666")
    meta.alignment = left
    ws_id.row_dimensions[2].height = 20
    header_row = 3
    _style_header_row(ws_id, header_row, EXCEL_COLUMNS)
    _fill_item_rows(ws_id, items, EXCEL_COLUMNS, header_row=header_row)
    ws_id.freeze_panes = "C4"

    # compact by-id
    compact_cols = [
        ("Id", "道具ID", 12),
        ("Secretname", "显示名", 28),
        ("Name", "分类/内部名", 16),
        ("Label", "标签", 10),
        ("Type", "类型Type", 10),
        ("Level", "等级", 8),
        ("Imagenumber", "图标编号", 10),
        ("BatchUse", "批量使用", 10),
    ]
    ws_compact = wb.create_sheet("按ID(简表)")
    ws_compact.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(compact_cols))
    t3 = ws_compact.cell(row=1, column=1, value="简表：仅 ID / 名称（完整属性见「详细属性」）")
    t3.font = Font(name="微软雅黑", bold=True, size=14, color="1F4E79")
    t3.alignment = left
    ws_compact.row_dimensions[1].height = 28
    meta3 = ws_compact.cell(row=2, column=1, value=f"条目数：{len(items)}")
    ws_compact.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(compact_cols))
    meta3.font = Font(name="微软雅黑", size=9, color="666666")
    _style_header_row(ws_compact, header_row, compact_cols)
    _fill_item_rows(ws_compact, items, compact_cols, header_row=header_row)
    ws_compact.freeze_panes = "A4"

    # Sheet: by display name
    by_name = sorted(items, key=lambda x: (str(x.get("Secretname") or ""), int(x.get("Id") or 0)))
    ws_name = wb.create_sheet("按名称")
    ws_name.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(NAME_INDEX_COLUMNS))
    title2 = ws_name.cell(row=1, column=1, value="按显示名排序（含攻防/特效摘要）")
    title2.font = Font(name="微软雅黑", bold=True, size=14, color="1F4E79")
    title2.alignment = left
    ws_name.row_dimensions[1].height = 28
    meta2 = ws_name.cell(
        row=2,
        column=1,
        value="提示：Ctrl+F 搜「生命之华」等；完整列请看「详细属性」",
    )
    ws_name.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(NAME_INDEX_COLUMNS))
    meta2.font = Font(name="微软雅黑", size=9, color="666666")
    meta2.alignment = left
    ws_name.row_dimensions[2].height = 20
    _style_header_row(ws_name, header_row, NAME_INDEX_COLUMNS)
    _fill_item_rows(ws_name, by_name, NAME_INDEX_COLUMNS, header_row=header_row)
    ws_name.freeze_panes = "A4"

    wb.save(path)


def write_csv(items: list[dict], path: Path) -> None:
    fields = [key for key, _, _ in EXCEL_COLUMNS]
    with path.open("w", encoding="utf-8-sig", newline="") as f:
        w = csv.DictWriter(f, fieldnames=fields, quoting=csv.QUOTE_ALL)
        w.writeheader()
        w.writerows(items)


def gm_search(items: list[dict], query: str) -> list[dict]:
    """Same logic as GMStorePanel.OnClickSearchCallBack."""
    q = query.strip()
    if not q:
        return items
    out = []
    for it in items:
        if q in it["Secretname"] or q in str(it["Id"]):
            out.append(it)
    return out


def main() -> int:
    parser = argparse.ArgumentParser(description="Export item_tbitemconfig (GM store table)")
    parser.add_argument("--bundle", type=Path, default=DEFAULT_BUNDLE)
    parser.add_argument(
        "--out",
        type=Path,
        default=DEFAULT_OUT,
        help="output path (.xlsx default, .csv if suffix is .csv)",
    )
    parser.add_argument("--json", type=Path, default=None, help="optional JSON output")
    parser.add_argument("--search", type=str, default=None, help="test GM local search")
    args = parser.parse_args()

    if not args.bundle.is_file():
        print(f"bundle missing: {args.bundle}", file=sys.stderr)
        return 1

    raw = extract_textasset_bytes(args.bundle, "item_tbitemconfig")
    items = load_items(raw)
    if args.search:
        hits = gm_search(items, args.search)
        print(f"GM search '{args.search}': {len(hits)} hit(s)")
        for it in hits[:20]:
            print(f"  Id={it['Id']}  Secretname={it['Secretname']}  Label={it['Label']}")
        if len(hits) > 20:
            print(f"  ... and {len(hits) - 20} more")

    args.out.parent.mkdir(parents=True, exist_ok=True)
    if args.out.suffix.lower() == ".csv":
        write_csv(items, args.out)
    else:
        if args.out.suffix.lower() not in (".xlsx", ".xlsm"):
            args.out = args.out.with_suffix(".xlsx")
        write_excel(items, args.out)
    print(f"exported {len(items)} items -> {args.out}")

    if args.json:
        args.json.write_text(json.dumps(items, ensure_ascii=False, indent=2), encoding="utf-8")
        print(f"json -> {args.json}")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
