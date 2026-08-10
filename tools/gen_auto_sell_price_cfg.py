#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""从「可上架装备表单_默认定价.xlsx」生成自动上架定价配置。

输出：tools/seqchapter_auto_stall/seqchapter_auto_sell_prices.txt
格式（每行）：<道具ID> <D价> <C价> <B价> <A价> <S价>   # 与 SeqChapterAutoStall.LoadPriceTable 一致
以「装备表单」页 道具ID + D/C/B/A/S 五档默认价为准；改价只需改 Excel 后重跑本脚本。

部署：打补丁时由 apply_combo_patch 的自动上架步骤把该 txt 复制到
cg37_Data/assets/hotfixdata/seqchapter_auto_sell_prices.txt（DLL 运行时读取）。
"""
from __future__ import annotations

import sys
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError as exc:
    raise SystemExit("需要 openpyxl：pip install openpyxl") from exc

ROOT = Path(__file__).resolve().parents[1]
XLSX = ROOT / "可上架装备表单_默认定价.xlsx"
OUT_TXT = ROOT / "tools" / "seqchapter_auto_stall" / "seqchapter_auto_sell_prices.txt"

GRADES = ["D", "C", "B", "A", "S"]
# 「装备表单」页列（1 起始）：
# 1分类 2装备名称 3等级 4道具ID 5配方ID 6D价 7C价 8B价 9A价 10S价 11默认货币
COL_ITEM_ID = 4
COL_PRICE_FIRST = 6


def main() -> int:
    if not XLSX.exists():
        raise SystemExit(f"找不到表单 {XLSX}（请先运行 tools/gen_craftable_equip_form.py 生成）")

    wb = load_workbook(XLSX, data_only=True, read_only=False)
    if "装备表单" not in wb.sheetnames:
        raise SystemExit(f"{XLSX} 缺少「装备表单」工作表")

    ws = wb["装备表单"]
    header_row = 3  # 与生成脚本一致
    lines = ["# 自动上架定价配置：<道具ID> D C B A S", "# 由 tools/gen_auto_sell_price_cfg.py 从默认定价.xlsx 生成，改价请改 Excel 后重跑"]
    count = 0
    for row in ws.iter_rows(min_row=header_row + 1):
        item_id = row[COL_ITEM_ID - 1].value
        if item_id is None:
            continue
        try:
            item_id = int(item_id)
        except (TypeError, ValueError):
            continue

        prices = []
        ok = True
        for g in range(len(GRADES)):
            cell = row[COL_PRICE_FIRST - 1 + g].value
            try:
                prices.append(int(cell))
            except (TypeError, ValueError):
                ok = False
                break
        if not ok:
            print(f"  跳过 {item_id}（价格列不完整）")
            continue

        lines.append(f"{item_id} {' '.join(str(p) for p in prices)}")
        count += 1

    OUT_TXT.write_text("\n".join(lines) + "\n", encoding="utf-8")
    print(f"已生成 {OUT_TXT}：{count} 条定价")
    print("（打补丁时该文件会随自动上架 DLL 一起部署到 hotfixdata）")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
