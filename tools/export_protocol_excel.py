#!/usr/bin/env python3
"""Export protocol_catalog.json to Excel with Chinese annotations."""
from __future__ import annotations

import json
import re
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
JSON_PATH = ROOT / "tools" / "protocol_catalog.json"
OUT_XLSX = ROOT / "tools" / "协议目录.xlsx"

# --- 协议中文名（枚举后缀 → 中文）---
PROTO_SUFFIX_CN: dict[str, str] = {
    "WALK": "行走移动",
    "MAP_CHECK": "地图校验",
    "MAP_EVENT": "地图事件",
    "MAP": "地图信息",
    "CREATE_BATTLE": "创建战斗",
    "BATTLE_REWARD": "战斗奖励",
    "BATTLE_PVP_REWARD": "PVP战斗奖励",
    "CREATE_BATTLE_PVP": "创建PVP战斗",
    "ENVIR_ONMENT": "战场环境",
    "LOOK_BATTLE": "观战/查看战斗",
    "BATTLE_COMMAND": "战斗指令",
    "MISSIONEX": "任务扩展",
    "ITEM": "道具同步",
    "SERVER_INFO": "服务器信息",
    "ITEM_MOVE": "道具移动(旧)",
    "USE_ITEM": "使用道具",
    "PICKUP_ITEM": "拾取道具",
    "DROP_ITEM": "丢弃道具",
    "DROP_GOLD": "丢弃金币",
    "DROP_PET": "丢弃/放出宠物",
    "MOVE_ITEM": "移动/整理道具",
    "ITEM_RECIPE": "道具配方",
    "TALK": "NPC对话",
    "TOOL_TIP": "工具提示",
    "UPDATE_PLAYER": "更新玩家",
    "UPDATE_PET": "更新宠物",
    "UPDATE_OBJ": "更新地图对象",
    "UPDATE_TEAM": "更新队伍",
    "OBJ_ADDITIONAL": "对象附加信息",
    "CLEAR_OBJ": "清除对象",
    "PLAYER_MAGIC": "玩家魔法",
    "UPDATE_TECH": "更新技能",
    "PET_TECH": "宠物技能",
    "CONFIG": "配置",
    "OPERATION_TEAM": "队伍操作",
    "OPERATION_PET": "宠物操作",
    "FRONT_CHAR_LIST": "前台角色列表",
    "ACTION": "角色动作",
    "ACTION_SKILL": "技能动作",
    "USE_TECH": "使用技能",
    "SKILL_EXCHANGE": "技能交换",
    "POS": "位置同步",
    "CHANGE_PET_NAME": "修改宠物名",
    "WINDOWS": "窗口消息",
    "MAP_EFFECT": "地图特效",
    "PALETTE": "调色板",
    "PLAYSE": "播放音效",
    "MAP_NAME": "地图名称",
    "DUNGEON": "迷宫",
    "LOGIN": "账号登录",
    "CREATECHAR": "创建角色",
    "CHARLOGIN": "角色登录",
    "CHARLOGOUT": "角色登出",
    "ECHO": "心跳回显",
    "LOGINGATE": "登录网关",
    "LOOK_NPC": "查看NPC",
    "NEWMESSAGE": "新消息",
    "RETURNBATTLE": "返回战斗",
    "MENU": "菜单/NPC选项",
    "RUN": "奔跑/快捷操作",
    "RED_DOT": "红点",
    "SELECTPLAYERNO": "选择角色编号",
    "VERIFY": "验证",
    "BATTLE_COMMAND_PLAYER": "战斗指令-玩家",
    "BATTLE_COMMAND_CHAR": "战斗指令-角色",
    "BATTLE_COMMAND_STATUS": "战斗指令-状态",
    "BATTLE_COMMAND_EXIT": "战斗指令-退出",
    "BATTLE_COMMAND_ACTION": "战斗指令-动作",
    "BATTLE_CONFIG": "自动战斗配置",
    "BATTLE_END": "战斗结束",
    "DELAY_TEST": "延迟测试",
    "DELAY_TEST_GATE": "网关延迟测试",
    "MISC": "杂项",
    "TALKEX": "扩展聊天/GM",
    "FRIEND": "好友",
    "FRIEND_CHANGE_OR_ADD": "好友增改",
    "FRIEND_DEL": "删除好友",
    "FRIEND_BLACK_DEL": "移除黑名单",
    "MAIL": "邮件",
    "MAIL_CHANGE_OR_ADD": "邮件增改",
    "MAIL_OPERATION": "邮件操作",
    "AREA": "采集区域",
    "RESET_TASK": "重置任务",
    "RECIPE": "配方/镶嵌",
    "PRODUCTION": "生产/制造",
    "TEAM": "组队",
    "STREET": "摆摊",
    "FRAME": "头像框",
    "PETBOOK": "宠物图鉴",
    "TRADE": "交易",
    "OFFLINECOMMAND": "离线指令",
    "TITLEEX": "称号",
    "HPFP": "血魔池",
    "PLAYERBUFF": "玩家BUFF",
    "ACCOMPANY": "佣兵",
    "ALLTALK": "全服聊天",
    "SELECTOBJ": "选择对象",
    "FAMILY": "家族",
    "VIGOR": "活力",
    "CRYSTAL": "水晶/抽奖",
    "DAILY_ACTIVITY": "日常活动",
    "TRUST": "悬赏/信任",
    "MULTI": "多控",
    "TEAMCONFIG": "队伍配置",
    "MESSAGE_CODE": "验证码",
    "NEARBY": "附近玩家",
    "CHARACTER_CHANGE": "角色变更",
    "OTHER_CODE": "其他验证码",
    "CUSTOMER": "客服",
    "LUCK": "幸运",
    "TASK_EXPAND": "任务扩展",
    "SHOP": "商店",
    "BATTLE_LOOK_PLAYER": "战斗观战玩家",
    "WATCH_BATTLE": "观战",
    "DAILY_EXPAND": "日常扩展",
    "PLAYER_SKILLS": "玩家技能栏",
    "PK": "PK",
    "WATCH": "观察",
    "QUERCHARACTER": "查询角色",
    "PLAYER_CONFIG": "玩家配置",
    "ACTIVITY": "活动",
    "PLAYER_POING": "角色加点",
    "MESSAGEBOX": "消息框",
    "PET_RESET": "宠物重置",
    "PET_POINT": "宠物加点",
    "DAILY_ACTIVITY_EX": "日常活动扩展",
    "ITEM_USE_EFFECT": "道具使用效果",
    "SKILL": "技能",
    "BACKPACK": "背包",
    "BANK": "银行/仓库",
    "ACCOUNT_BANK": "账号仓库",
    "BATTLE_SPEED": "战斗速度",
    "AUCTIONHOUSE": "拍卖行",
    "RED_POINT": "红点",
    "ITEM_LIST": "道具列表",
    "FIX_POS": "修正坐标",
    "LOCK": "安全锁",
    "FIX_DUNGEON": "固定副本",
    "MONSTER_BREED": "魔物培育",
    "PET_UPSTAR": "宠物升星",
    "LAYERBOX": "层数宝箱",
    "REPAIRBATTLE": "修复战斗",
    "CLIENTLIMIT": "客户端限制",
    "EASTEREGG": "彩蛋",
    "AUTOBATTLE": "自动战斗",
    "JOBSWITCH": "职业切换",
    "FLASHSALE": "闪购",
    "PETBOND": "宠物羁绊",
    "FLASHLIMIT": "限时闪购",
    "SELECT_PET_WINDOWS": "选择宠物窗口",
    "CODE_PET_WINDOWS": "宠物码窗口",
    "FIVE_CHAR_DRAW": "五字抽奖",
    "PET_EQUIP": "宠物装备",
    "EQUIP_FORCE": "装备强化",
    "DUR_REPAIRE": "耐久修理",
    "FAMILY_BATTLE": "家族战",
    "CHOOSE_PET": "选择宠物",
    "BLINDBOX": "盲盒",
    "PET_RIDE": "宠物骑乘",
    "OTHER_EQUIP": "其他装备",
    "PET_RESETBASE": "宠物洗点",
    "PRAY": "祈祷",
    "MONSTERTOWER": "魔物塔",
    "CURRENCY_EXCHANGE": "货币兑换",
    "MAP_UPDATE": "地图更新",
    "ONLINE_INFO": "线路/在线信息",
    "TASK_WARP": "任务传送",
    "FLURRYPK_BATTLE": "乱斗PK",
    "FLURRYPK_RANK": "乱斗PK排行",
    "SET_SPECIAL_FLAG": "特殊标记",
    "USE_CURRENCY_TIP": "货币提示",
    "SKIN_LEVELUP": "皮肤升级",
    "CRYSTAL_HOUSE": "水晶屋",
    "SOLDIER_OF_HONOR": "荣誉士兵",
    "BOSS_SUPPRESS_TOKEN": "BOSS讨伐令",
    "EARCH_MOUSE_LOTTERY": "地鼠抽奖",
    "HERO_TRIALS": "英雄试炼",
    "WORLD_BOSS": "世界BOSS",
    "BLIND_BOX_NEW1": "新盲盒",
    "BOSS_LAND_CHALLENGE": "BOSS大陆挑战",
    "ENCHANT_FUNCTION": "附魔",
    "PLAYER": "玩家管理",
    "RANKLIST": "排行榜",
    "PET": "宠物",
    "LEVEL": "等级",
    "KICK_OUT": "踢出",
    "PAY_LUA": "充值Lua",
    "RISK_CONTROL": "风控",
    "PLAYER_EFFECT": "玩家特效",
    "NEWCOMER_7DAY": "新手7日",
    "DUNGEON_ENTER": "进入副本",
    "PET_SACRIFICE": "宠物献祭",
    "BOX_FRUITY": "水果宝箱",
    "GEM_FUNCTION": "宝石功能",
    "SELECT_BOSS_BOX": "选择BOSS箱",
    "FAMILY_TRIAL": "家族试炼",
    "FAMILY_BRAWL": "家族乱斗",
    "FAMILY_REDPACK": "家族红包",
    "NEWYEARSDAY2025": "2025元旦",
    "BATTLE_UPLOAD": "战斗上传",
    "LOGIN_SWITCH": "登录切换",
    "BRAWL_RANK": "乱斗排行",
    "BROADCAST_ROOM": "广播房间",
    "LEGACYOFTHELOST": "失落遗产",
    "TRADE_BANK": "交易银行",
    "LUCK_ROCKETS": "幸运火箭",
    "PLAYER_ACOUSTIC": "玩家声学",
    "PET_POTENTIAL": "宠物潜力",
    "SHOVEL_TREASURE": "铲子挖宝",
    "OFFLINE_TRADE": "离线交易",
    "HUNDRED_LAYER_BATTLE": "百层战斗",
    "PET_MAX_CREST_EFFECT": "宠物满纹效果",
    "LOOPY_TRIAL": "露比试炼",
    "PET_EXCHANGE_OR_MERGE": "宠物兑换合并",
    "JEWELRY_MERGE": "首饰合并",
    "PET_REFORM_MUTATION": "宠物改造",
    "MONTH_CARD_PRIVILEGE": "月卡特权",
    "BATTLE_PASS_ADVENTURE": "战令冒险",
    "PET_REFACTOR_DESTRUCT": "宠物重构销毁",
    "PET_RECYCLE_EXCHANGE": "宠物回收兑换",
    "DAILY_PASS": "每日通行证",
    "SKY_DORP_RED_PACK": "天降红包",
    "SERVER_HEARTBEAT": "服务端心跳",
    "SERVER_OTHERMSG": "服务端其他消息",
    "SERVER_ENCRYPTIONDATA": "服务端加密",
    "SERVER_BROADCASTENCRYPTIONDATA": "服务端广播加密",
    "SERVER_FD_CLOSE": "服务端关闭连接",
    "SERVER_DECRYPTIONDATA": "服务端解密",
    "SERVER_BROADCASTDECRYPTIONDATA": "服务端广播解密",
}

# --- 字段中文注释 ---
FIELD_CN: dict[str, str] = {
    "Type": "操作类型（字符串，区分具体功能）",
    "KUid": "角色唯一ID（必填，多控时指定目标角色）",
    "Id": "编号/索引/目标ID",
    "Num": "数量",
    "Index": "索引/格子号",
    "Haveitemindex": "背包格子索引（-1表示无）",
    "Haveitemindex2": "背包格子索引2",
    "Haveitemindex3": "背包格子索引3",
    "Flg": "标志位/页签/开关",
    "X": "地图X坐标",
    "Y": "地图Y坐标",
    "X1": "区域X1",
    "Y1": "区域Y1",
    "X2": "区域X2",
    "Y2": "区域Y2",
    "Direction": "朝向",
    "Dir": "方向",
    "Mapid": "地图ID",
    "Floor": "楼层/层数",
    "Event": "事件类型",
    "Seqno": "事件序号",
    "Result": "结果码",
    "Str": "指令字符串/文本内容",
    "Count": "计数/批次",
    "Isauto": "是否自动战斗",
    "BattleTime": "战斗剩余时间",
    "Battleindex": "战斗索引",
    "Uid": "用户/对象UID",
    "Channel": "聊天频道",
    "Msg": "消息内容",
    "Toid": "私聊目标UID",
    "Message": "对话消息",
    "Cdkey": "账号",
    "Passwd": "密码",
    "Version": "客户端版本",
    "Plat": "平台",
    "DeviceId": "设备ID",
    "Device": "设备类型",
    "ClientIp": "客户端IP",
    "DeviceInfo": "设备信息",
    "ServerId": "服务器ID",
    "DeviceName": "设备名称",
    "Name": "名称",
    "Objindex": "地图对象索引",
    "Fromindex": "源背包格",
    "Toindex": "目标背包格",
    "Selectindex": "选择索引",
    "Usecount": "使用数量",
    "Itemindex": "道具索引",
    "ActivityId": "活动ID",
    "Code": "兑换码/验证码",
    "IndexList": "索引列表（批量操作）",
    "CurrencyType": "货币类型（PROTO_CURRENCY枚举）",
    "ExchangeAmount": "兑换数量",
    "ExchangedCurrencyType": "兑换后货币类型",
    "ExchangedCurrencyAmount": "兑换后数量",
    "TargetCurrencyType": "目标货币类型",
    "Rate": "比率/阈值",
    "RateDescribe": "比率描述",
    "ItemId": "道具ID",
    "BagType": "背包类型",
    "Pile": "堆叠数",
    "Time": "时间",
    "Hp": "生命值",
    "Fp": "魔法值/魔力",
    "CurHp": "当前HP",
    "CurFp": "当前FP",
    "Config": "配置对象",
    "Players": "多控角色列表",
    "TabIndex": "商店页签",
    "Curr": "货币类型/当前项",
    "BoxIndexs": "盲盒索引列表",
    "ClientIndice": "客户端选中索引",
    "TeamName": "队伍名称",
    "Shopid": "商店ID",
    "PVPShopNum": "PVP商店数量",
    "DrawCount": "抽奖次数",
    "DrawSelect": "抽奖选择",
    "TargetX": "目标X",
    "TargetY": "目标Y",
    "Pid": "池ID/宠物池",
    "Tid": "模板ID/宠物ID",
    "LineId": "线路ID",
    "WarpId": "传送点ID",
    "Vital": "体力",
    "Str": "力量/指令串",
    "Tgh": "强度",
    "Quick": "速度",
    "Magic": "魔法",
    "Bpindex": "BP方案索引",
    "Earth": "地属性",
    "Water": "水属性",
    "Fire": "火属性",
    "Wind": "风属性",
    "ItemCostType": "材料来源（0背包/1其他）",
    "MakeStoreType": "产物去向（0背包/1其他）",
    "QualityItemIndex": "品质道具格",
    "FirstGoodIndex": "首优道具格",
    "GetQuality": "获取品质",
    "Level": "等级",
    "Exp": "经验",
    "Nexpexp": "下级经验",
    "Gold": "金币",
    "Status": "状态",
    "Rank": "排名",
    "Aid": "辅助ID",
    "Round": "回合数",
    "BpFlg": "BP标志",
    "ItemLimit": "道具限制",
    "PlayerSkillFlag": "玩家技能标志",
    "PetSkillFlag": "宠物技能标志",
    "TransformState": "变身状态",
    "MaxRound": "最大回合",
    "LoopyTrialLevel": "露比试炼层数",
}

# --- Type 操作中文说明（常见）---
TYPE_CN: dict[str, str] = {
    "获取数据": "拉取面板/仓库数据",
    "开始采集": "开始自动采集（Flg=页签+1）",
    "停止采集": "停止采集",
    "放入需要道具": "提交采集所需材料",
    "删除物品": "删除采集仓库物品",
    "取出物品": "取出到个人背包",
    "取出物品到账号仓库": "取出到账号仓库",
    "获取多控": "拉取多控角色列表",
    "登陆角色": "多控登录子角色",
    "召唤角色": "召唤子角色到当前地图",
    "切换角色": "切换主控角色",
    "登出角色": "多控登出子角色",
    "一键召唤": "一键召唤所有子角色",
    "头像切换角色": "从头像栏切换角色",
    "血魔池状态": "查询血池/魔池状态",
    "血魔池设置": "打开血魔池设置面板",
    "开启血池": "开启自动回血池",
    "关闭血池": "关闭血池",
    "开启魔池": "开启自动回魔池",
    "关闭魔池": "关闭魔池",
    "获取血池道具": "从血池取道具",
    "获取魔池道具": "从魔池取道具",
    "血池阈值": "设置血池触发阈值",
    "魔池阈值": "设置魔池触发阈值",
    "存道具": "仓库存入道具",
    "取道具": "仓库取出道具",
    "整理仓库": "整理仓库排序",
    "存宠物": "仓库存入宠物",
    "取宠物": "仓库取出宠物",
    "个人仓库": "打开个人仓库",
    "获取信息": "拉取兑换信息",
    "兑换货币": "执行货币兑换（如工时）",
    "镶嵌宝石": "装备镶嵌宝石",
    "制造道具": "配方制造道具",
    "活动信息": "拉取指定活动详情",
    "活动列表": "拉取活动列表",
    "CDKey兑换": "CDKey礼包兑换",
    "giftCode": "礼包码兑换",
    "每日签到": "活动每日签到",
    "远程个人仓库": "远程打开个人仓库",
    "远程账号道具仓库": "远程打开账号道具仓",
    "远程个人道具仓库": "远程打开个人道具仓",
    "远程个人宠物仓库": "远程打开个人宠物仓",
    "远程出售魔石": "远程出售魔石",
    "玩家BUFF数据": "拉取玩家BUFF信息",
    "自动战斗": "自动战斗相关",
    "标记": "战斗标记设置",
    "自动技能设置": "自动技能配置",
    "宠物改造": "宠物改造操作",
    "替换改造": "替换改造结果",
    "商店信息": "拉取商店数据",
    "购买商品": "购买商店商品",
    "背包信息": "拉取背包数据",
    "创建队伍": "创建新队伍",
    "加入队伍": "加入队伍",
    "离开队伍": "离开队伍",
    "解散队伍": "解散队伍",
    "踢出队伍": "踢出队员",
    "转让队伍": "转让队长",
    "比赛报名": "PK比赛报名",
    "同步数据": "同步试炼数据",
    "领取奖励": "领取奖励",
    "创建试炼队伍": "露比试炼创建队伍",
}

GM_TYPE_CN: dict[str, str] = {
    "clearplayerbag": "GM命令：清空背包",
    "delpet all": "GM命令：删除全部宠物",
    "delstreet ": "GM命令：删除摊位",
    "superman": "GM命令：超级模式",
    "warp 0 ": "GM命令：传送地图0",
    "warp 1 ": "GM命令：传送地图1",
}

TYPE_PREFIX_CN: list[tuple[str, str]] = [
    ("获取", "查询/拉取数据"),
    ("领取", "领取奖励或物品"),
    ("远程", "远程功能（通常 ActivityId=19）"),
    ("购买", "商店购买"),
    ("打开", "打开界面/面板"),
    ("开始", "开始执行"),
    ("停止", "停止执行"),
    ("关闭", "关闭功能"),
    ("开启", "开启功能"),
    ("删除", "删除操作"),
    ("设置", "设置参数"),
    ("修改", "修改信息"),
    ("整理", "整理/排序"),
    ("装备", "穿戴装备或外观"),
    ("卸下", "卸下装备或外观"),
    ("一键", "批量快捷操作"),
    ("请求", "请求服务端数据"),
    ("提交", "提交任务/物品"),
    ("重置", "重置状态"),
    ("使用", "使用物品/功能"),
    ("加入", "加入组织/队伍"),
    ("离开", "离开组织/队伍"),
    ("解散", "解散组织/队伍"),
    ("踢出", "踢出成员"),
    ("转让", "转让权限"),
    ("兑换", "兑换物品/货币"),
    ("上架", "上架出售"),
    ("下架", "下架商品"),
    ("确认", "确认操作"),
    ("取消", "取消操作"),
    ("搜索", "搜索查询"),
    ("发送", "发送消息/邀请"),
    ("添加", "添加好友等"),
    ("拉黑", "拉黑好友"),
    ("同意", "同意申请"),
    ("邀请", "发送邀请"),
    ("挑战", "发起挑战"),
    ("捐献", "家族/公会捐献"),
    ("升级", "升级强化"),
    ("修复", "修理装备"),
    ("鉴定", "鉴定道具"),
    ("拆分", "拆分道具堆叠"),
    ("合并", "合并道具"),
    ("注入", "注入材料"),
    ("抽奖", "抽奖操作"),
    ("许愿", "许愿活动"),
    ("签到", "签到领奖"),
    ("挂机", "挂机相关"),
    ("观战", "观战相关"),
    ("交易", "玩家交易"),
    ("仓库", "仓库操作"),
    ("背包", "背包操作"),
    ("宠物", "宠物相关"),
    ("家族", "家族/公会"),
    ("队伍", "组队相关"),
    ("多控", "多角色控制"),
    ("血魔", "血魔池"),
    ("血池", "血池"),
    ("魔池", "魔池"),
    ("采集", "采集区域"),
    ("活动", "活动系统"),
    ("战令", "战令/通行证"),
    ("月卡", "月卡特权"),
    ("皮肤", "皮肤外观"),
    ("头饰", "头饰外观"),
    ("翅膀", "翅膀外观"),
    ("称号", "称号系统"),
    ("头像", "头像框"),
    ("聊天", "聊天框外观"),
    ("骑宠", "骑宠皮肤"),
    ("附魔", "装备附魔"),
    ("宝石", "宝石镶嵌/升阶"),
    ("首饰", "首饰合并/维修"),
    ("历练", "历练系统"),
    ("加点", "属性加点"),
    ("切换", "切换方案/角色"),
    ("开通", "开通功能栏位"),
    ("验证", "验证密码/手机"),
    ("标记", "标记设置"),
    ("加速", "战斗加速"),
    ("掉落", "活动掉落开关"),
    ("分享", "分享/排行榜"),
    ("批量", "批量操作"),
    ("全部", "全部取出/操作"),
    ("指定", "指定目标"),
    ("瞬间", "瞬间移动"),
    ("手机", "手机验证"),
    ("短信", "短信验证"),
    ("NPC", "NPC服务"),
    ("PK", "PK切磋"),
]

HEADER_FILL = PatternFill("solid", fgColor="4472C4")
HEADER_FONT = Font(color="FFFFFF", bold=True)
ALT_FILL = PatternFill("solid", fgColor="F2F2F2")


def proto_name_cn(enum_name: str) -> str:
    core = enum_name.replace("LSSPROTO_", "").replace("_FUNC", "")
    if core in PROTO_SUFFIX_CN:
        return PROTO_SUFFIX_CN[core]
    # try strip LSSPROTO_ prefix only
    for key, val in PROTO_SUFFIX_CN.items():
        if core == key or core.endswith(key):
            return val
    # fallback: humanize
    return core.replace("_", " ").lower()


def field_comment(name: str, proto_class: str = "") -> str:
    if name in FIELD_CN:
        return FIELD_CN[name]
    # heuristic
    low = name.lower()
    if low.endswith("index"):
        return "索引/格子号"
    if low.endswith("id"):
        return "ID"
    if low.endswith("num") or low.endswith("count"):
        return "数量"
    if low.startswith("is"):
        return "布尔标志"
    if name == "Info":
        return "信息对象/详情"
    if name == "Data1" or name == "Data2":
        return "扩展数据"
    if name == "Item":
        return "道具数据"
    return ""


def type_comment(type_str: str, opcode_name: str = "") -> str:
    if type_str in TYPE_CN:
        return TYPE_CN[type_str]
    if type_str in GM_TYPE_CN:
        return GM_TYPE_CN[type_str]
    # battle command codes
    battle_map = {
        "@": "战斗指令：攻击",
        "I": "战斗指令：防御",
        "E": "战斗指令：逃跑",
        "W": "战斗指令：换位",
        "G": "战斗指令：捕获",
        "P": "战斗指令：宠物",
        "S": "战斗指令：技能",
        "H": "战斗指令：使用物品",
        "N": "战斗指令：待机",
        "U": "战斗指令：特殊",
        "M": "战斗指令：魔法",
    }
    if opcode_name == "LSSPROTO_BATTLE_COMMAND_FUNC":
        for k, v in battle_map.items():
            if type_str.startswith(k):
                return v
    # already Chinese — prefix heuristic
    if re.search(r"[\u4e00-\u9fff]", type_str):
        for prefix, desc in TYPE_PREFIX_CN:
            if type_str.startswith(prefix) or prefix in type_str[:4]:
                return f"{desc}（{type_str}）"
        return f"游戏内操作（{type_str}）"
    # English fallback
    if type_str.isascii():
        return f"英文字符串操作（{type_str}）"
    return ""


def style_header(ws, row: int = 1) -> None:
    for cell in ws[row]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def auto_width(ws, min_w: int = 10, max_w: int = 48) -> None:
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        width = min_w
        for cell in col:
            if cell.value:
                width = max(width, min(len(str(cell.value)) + 2, max_w))
        ws.column_dimensions[letter].width = width


def sheet_protocol_overview(wb: Workbook, data: dict) -> None:
    ws = wb.active
    ws.title = "协议总览"
    headers = [
        "协议编号",
        "协议枚举名",
        "中文名称",
        "客户端发包",
        "发包次数",
        "Proto_CS类型",
        "Type操作数",
        "说明",
    ]
    ws.append(headers)
    protocols = data["protocols"]
    for name in sorted(protocols, key=lambda k: protocols[k]["opcode"]):
        p = protocols[name]
        cn = proto_name_cn(name)
        client = "是" if p["client_send"] else "否"
        protos = ", ".join(p.get("proto_types") or []) or "—"
        tc = len(p.get("type_literals") or [])
        note = "服务端下发/未使用" if not p["client_send"] else ""
        if p["opcode"] >= 10_000_000:
            note = "服务端专用协议"
        ws.append(
            [
                p["opcode"],
                name,
                cn,
                client,
                p["send_count"] or "",
                protos,
                tc if tc else "",
                note,
            ]
        )
    style_header(ws)
    ws.freeze_panes = "A2"
    auto_width(ws)


def sheet_type_operations(wb: Workbook, data: dict) -> None:
    ws = wb.create_sheet("Type操作名")
    headers = [
        "协议编号",
        "协议枚举名",
        "协议中文名",
        "Type/操作名",
        "中文说明",
        "关联Proto",
        "Send封装方法",
    ]
    ws.append(headers)
    protocols = data["protocols"]
    for name in sorted(protocols, key=lambda k: protocols[k]["opcode"]):
        p = protocols[name]
        if not p.get("type_literals"):
            continue
        cn = proto_name_cn(name)
        methods = ", ".join(w["method"] for w in p.get("send_wrappers") or [])
        protos = ", ".join(p.get("proto_types") or [])
        for t in p["type_literals"]:
            ws.append(
                [
                    p["opcode"],
                    name,
                    cn,
                    t,
                    type_comment(t, name),
                    protos,
                    methods,
                ]
            )
    style_header(ws)
    ws.freeze_panes = "A2"
    auto_width(ws)


def sheet_proto_fields(wb: Workbook, data: dict) -> None:
    ws = wb.create_sheet("Proto字段")
    headers = [
        "Proto_CS类名",
        "字段名",
        "字段中文注释",
        "关联协议编号",
        "关联协议名",
        "协议中文名",
    ]
    ws.append(headers)
    proto_classes = data.get("proto_classes") or {}
    # reverse map proto -> opcodes
    proto_to_ops: dict[str, list[tuple[int, str]]] = {}
    for pname, p in data["protocols"].items():
        for pt in p.get("proto_types") or []:
            proto_to_ops.setdefault(pt, []).append((p["opcode"], pname))

    for cls in sorted(proto_classes):
        fields = proto_classes[cls]
        ops = proto_to_ops.get(cls, [])
        op_str = ", ".join(str(o[0]) for o in ops) if ops else ""
        op_names = ", ".join(o[1] for o in ops) if ops else ""
        op_cn = ", ".join(proto_name_cn(o[1]) for o in ops) if ops else ""
        for i, field in enumerate(fields):
            ws.append(
                [
                    cls if i == 0 else "",
                    field,
                    field_comment(field, cls),
                    op_str if i == 0 else "",
                    op_names if i == 0 else "",
                    op_cn if i == 0 else "",
                ]
            )
    style_header(ws)
    ws.freeze_panes = "A2"
    auto_width(ws)


def sheet_send_wrappers(wb: Workbook, data: dict) -> None:
    ws = wb.create_sheet("Send封装")
    headers = [
        "协议编号",
        "协议枚举名",
        "协议中文名",
        "方法名",
        "方法签名",
        "Proto类型",
        "Type参数来源",
        "字段赋值说明",
        "方法内Type常量",
    ]
    ws.append(headers)
    protocols = data["protocols"]
    for name in sorted(protocols, key=lambda k: protocols[k]["opcode"]):
        p = protocols[name]
        if not p.get("send_wrappers"):
            continue
        cn = proto_name_cn(name)
        for w in p["send_wrappers"]:
            assigns = w.get("field_assigns") or {}
            assign_str = "; ".join(f"{k}={v[0]}" for k, v in assigns.items()) if assigns else ""
            lits = ", ".join(w.get("type_literals_in_body") or [])
            ws.append(
                [
                    p["opcode"],
                    name,
                    cn,
                    w["method"],
                    f"{w['method']}({w['args']})",
                    w.get("proto_type") or "",
                    w.get("type_source") or "",
                    assign_str,
                    lits,
                ]
            )
    style_header(ws)
    ws.freeze_panes = "A2"
    auto_width(ws, max_w=60)


def sheet_field_glossary(wb: Workbook) -> None:
    ws = wb.create_sheet("字段注释词典")
    headers = ["字段名", "中文注释", "备注"]
    ws.append(headers)
    notes = {
        "KUid": "几乎所有业务协议都需要，助手 send_proto 可自动补全",
        "Type": "多态协议的核心字段，字符串区分操作",
        "CurrencyType": "见 PROTO_CURRENCY 枚举，如 PROTO_CURRENCY_AREA_TIME=工时",
        "Haveitemindex": "背包格子，从0开始，-1表示未指定",
        "IndexList": "批量存取道具时使用",
    }
    for field in sorted(FIELD_CN):
        ws.append([field, FIELD_CN[field], notes.get(field, "")])
    style_header(ws)
    ws.freeze_panes = "A2"
    auto_width(ws)


def sheet_usage_guide(wb: Workbook) -> None:
    ws = wb.create_sheet("使用说明")
    lines = [
        ["魔力宝贝：序章 — 客户端发送协议 Excel 目录"],
        [""],
        ["数据来源", "tools/hotfix_ilspy/hotfix.dll.bytes.decompiled.cs"],
        ["生成脚本", "tools/export_protocol_excel.py"],
        ["JSON源", "tools/protocol_catalog.json"],
        [""],
        ["工作表说明"],
        ["协议总览", "229条 LSSPROTO 协议索引，含编号、中文名、是否发包、Proto类型"],
        ["Type操作名", "带 Type 字段的协议，列出所有操作名字符串及中文说明"],
        ["Proto字段", "171个 Proto_CS_* 类的 protobuf 字段及中文注释"],
        ["Send封装", "游戏内 Send* 封装方法、参数签名、字段赋值关系"],
        ["字段注释词典", "常用字段中文对照"],
        [""],
        ["助手发包示例"],
        ["通用", 'send_proto(id, opcode=1008, proto_type="Proto_CS_Area", fields={"Type":"获取数据","KUid":uid})'],
        ["GM命令", 'send_gm(id, "additem 1 1001 99")  → 协议1000 TALKEX'],
        ["使用道具", "send_use_item(id, haveitemindex=0, uid=uid, x=0, y=0)  → 协议16"],
        ["账号仓库存", "send_account_bank_deposit(id, haveitemindex=0, uid=uid)  → 协议2003"],
        ["镶嵌宝石", "send_gem_inlay(id, equip_index=0, gem_index=1, uid=uid)  → 协议1010"],
        [""],
        ["注意"],
        ["1", "编号≥10000000 为服务端下发协议，客户端不发送"],
        ["2", "1047 ACTIVITY 等同编号可能对应多个 Send* 方法，Type+ActivityId 组合区分"],
        ["3", "战斗指令 Str 字段为特殊编码，见 Type操作名 中 LSSPROTO_BATTLE_COMMAND_FUNC"],
    ]
    for row in lines:
        ws.append(row)
    ws["A1"].font = Font(bold=True, size=14)
    auto_width(ws, max_w=80)


def main() -> None:
    data = json.loads(JSON_PATH.read_text(encoding="utf-8"))
    wb = Workbook()
    sheet_protocol_overview(wb, data)
    sheet_type_operations(wb, data)
    sheet_proto_fields(wb, data)
    sheet_send_wrappers(wb, data)
    sheet_field_glossary(wb)
    sheet_usage_guide(wb)
    wb.save(OUT_XLSX)
    print(f"Wrote {OUT_XLSX} ({OUT_XLSX.stat().st_size // 1024} KB)")


if __name__ == "__main__":
    main()
