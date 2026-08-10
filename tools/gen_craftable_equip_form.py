#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""生成「可制造装备可上架表单」Excel。

数据来源（只读配置，均为 Luban ByteBuf）：
  - excelgeneral/other_tbrecipeinfodescconfig.bytes  制造配方（skillid<=214 为装备）
  - excelgeneral/item_tbitemconfig.bytes             道具表（装备名称/等级/品质等）
  - excelgeneral/item_tbequipscorelevelconfig.bytes  装备评分品级阈值（Id 1-5）

品级规则（与客户端 ItemManager 一致）：
  GetEquipScoreLevel：按属性实际值/满值归一化加权得分 → 百分比
  阈值 item_tbequipscorelevelconfig：Id1:10 / Id2:40 / Id3:80 / Id4:100 / Id5:999
  Id1=最低品级(D) … Id5=最高品级(S)，图标 equipscore{Id}.png。

定价规则（可上架表单默认值，首次配置时展示）：
  D 级 = 等级² × 1000（取整）
  C/B/A/S 每上一档 +10%（对 D 逐档累乘 1.1，取整）
  默认销售货币 = 金币（上架 Tcost=0）
"""
from __future__ import annotations

import io
import sys
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
except ImportError as exc:
    raise SystemExit("需要 openpyxl：pip install openpyxl") from exc

ROOT = Path(__file__).resolve().parents[1]
EXTRACT = ROOT / "tools" / "_config_extract" / "excelgeneral"
RECIPE_PATH = EXTRACT / "other_tbrecipeinfodescconfig.bytes"
ITEM_PATH = EXTRACT / "item_tbitemconfig.bytes"
SCORE_LEVEL_PATH = EXTRACT / "item_tbequipscorelevelconfig.bytes"
OUT_XLSX = ROOT / "可上架装备表单_默认定价.xlsx"

# 装备类型 Type -> 分类名（来自 item_tbitemtypeconfig：武器/防具）
EQUIP_TYPES = {
    0: "单手剑", 1: "斧头", 2: "枪", 3: "法杖", 4: "弓",
    5: "小刀", 6: "回力镖", 7: "盾",
    8: "头盔", 9: "帽子", 10: "铠甲", 11: "衣服", 12: "长袍",
    13: "长靴", 14: "鞋子",
}
EQUIP_TYPE_RANGE = range(0, 15)

# 品级档位（低 -> 高）
GRADES = ["D", "C", "B", "A", "S"]
GRADE_UP_RATE = 0.10  # 每上一品级 +10%


class ByteBuf:
    """Bright.Serialization.ByteBuf 精简读实现（与 export_item_list.py 一致）。"""

    __slots__ = ("_data", "_i")

    def __init__(self, data: bytes, offset: int = 0) -> None:
        self._data = data
        self._i = offset

    @property
    def position(self) -> int:
        return self._i

    def read_uint(self) -> int:
        data = self._data
        i = self._i
        num = data[i]
        if num < 128:
            self._i = i + 1
            return num
        if num < 192:
            self._i = i + 2
            return ((num & 0x3F) << 8) | data[i + 1]
        if num < 224:
            self._i = i + 3
            return ((num & 0x1F) << 16) | (data[i + 1] << 8) | data[i + 2]
        if num < 240:
            self._i = i + 4
            return (
                ((num & 0xF) << 24)
                | (data[i + 1] << 16)
                | (data[i + 2] << 8)
                | data[i + 3]
            )
        self._i = i + 5
        return (
            (data[i + 1] << 24)
            | (data[i + 2] << 16)
            | (data[i + 3] << 8)
            | data[i + 4]
        ) & 0xFFFFFFFF

    def read_int(self) -> int:
        u = self.read_uint()
        return u - 0x100000000 if u >= 0x80000000 else u

    def read_size(self) -> int:
        return self.read_uint()

    def read_string(self) -> str:
        n = self.read_size()
        if n <= 0:
            return ""
        end = self._i + n
        if end > len(self._data):
            raise ValueError(f"string overruns buffer at {self._i}, n={n}")
        s = self._data[self._i : end].decode("utf-8")
        self._i = end
        return s

    def read_bool(self) -> bool:
        v = self._data[self._i]
        self._i += 1
        return v != 0


_INT_FIELDS_AFTER_LABEL = [
    "BatchUse", "Id", "Imagenumber", "Cost", "Type", "Quality", "Rank", "Bothhand",
    "Fieldtype", "Ableusebattle", "Target", "Maxremain", "Level", "Basefailedprob",
    "DurabilityMin", "DurabilityMax", "AttacknumMin", "AttacknumMax",
    "Ableeffectbetweenhave", "Modflg", "AttackMin", "AttackMax", "DefenceMin",
    "DefenceMax", "AgilityMin", "AgilityMax", "MagicMin", "MagicMax", "RecoveryMin",
    "RecoveryMax", "CriticalMin", "CriticalMax", "CounterMin", "CounterMax",
    "HitrateMin", "HitrateMax", "AvoidMin", "AvoidMax", "HpMin", "HpMax", "FpMin",
    "FpMax", "LuckMin", "LuckMax", "CharismaMin", "CharismaMax", "CharmMin",
    "CharmMax", "Attrib", "Attrib2", "Attribvalue", "Attribvalue2", "StaminaMin",
    "StaminaMax", "DexMin", "DexMax", "IntelligenceMin", "IntelligenceMax",
    "PoisonMin", "PoisonMax", "SleepMin", "SleepMax", "StoneMin", "StoneMax",
    "DrunkMin", "DrunkMax", "ConfusionMin", "ConfusionMax", "AmnesiaMin",
    "AmnesiaMax", "Specialeffect", "Specialeffectvalue", "Specialeffectvalue2",
    "MatWeapon", "MatArmour", "MatAccessory", "Useaction",
]


def _read_item(buf: ByteBuf) -> dict:
    fields = {
        "Name": buf.read_string(),
        "Secretname": buf.read_string(),
        "Label": buf.read_string(),
    }
    for key in _INT_FIELDS_AFTER_LABEL:
        fields[key] = buf.read_int()
    fields["Dropatlogout"] = buf.read_bool()
    fields["Vanishatdrop"] = buf.read_bool()
    fields["Canpetmail"] = buf.read_bool()
    fields["RssMin"] = buf.read_int()
    fields["RssMax"] = buf.read_int()
    fields["Cansell"] = buf.read_bool()
    for key in ("Exp1", "Exp2", "Rareflg", "Inboxflg", "AdmMin", "AdmMax", "Sellunit"):
        fields[key] = buf.read_int()
    n = buf.read_size()
    if n < 0 or n > 10000:
        raise ValueError(f"Resource list too large: {n}")
    fields["Resource"] = [buf.read_int() for _ in range(n)]
    return fields


def load_items() -> dict[int, dict]:
    buf = ByteBuf(io.open(ITEM_PATH, "rb").read())
    count = buf.read_size()
    if count <= 0 or count > 500_000:
        raise ValueError(f"suspicious item count: {count}")
    items: dict[int, dict] = {}
    for _ in range(count):
        row = _read_item(buf)
        items[int(row["Id"])] = row
    return items


def load_recipes() -> list[dict]:
    buf = ByteBuf(io.open(RECIPE_PATH, "rb").read())
    count = buf.read_size()
    recipes = []
    for _ in range(count):
        rid = buf.read_int()
        name = buf.read_string()
        makeid = buf.read_int()
        skillid = buf.read_int()
        exp = buf.read_int()
        ncost = buf.read_size()
        costs = []
        for _ in range(ncost):
            costs.append((buf.read_int(), buf.read_int(), buf.read_int()))
        typelv = buf.read_int()
        recipes.append(
            {
                "rid": rid,
                "name": name,
                "makeid": makeid,
                "skillid": skillid,
                "exp": exp,
                "costs": costs,
                "typelv": typelv,
            }
        )
    return recipes


def load_score_levels() -> dict[int, int]:
    buf = ByteBuf(io.open(SCORE_LEVEL_PATH, "rb").read())
    count = buf.read_size()
    levels = {}
    for _ in range(count):
        sid = buf.read_int()
        score = buf.read_int()
        levels[sid] = score
    return levels


def default_price(level: int, grade_idx: int) -> int:
    """D 级 = 等级²×1000；每上一品级 ×(1+10%)，取整。"""
    base = level * level * 1000
    for _ in range(grade_idx):
        base = int(base * (1 + GRADE_UP_RATE))
    return base


def main() -> int:
    items = load_items()
    recipes = load_recipes()
    score_levels = load_score_levels()

    equip_recipes = [r for r in recipes if r["skillid"] <= 214]
    rows = []
    for r in sorted(equip_recipes, key=lambda x: (x["skillid"], x["typelv"], x["rid"])):
        it = items.get(r["makeid"])
        if it is None:
            continue
        itype = int(it["Type"])
        if itype not in EQUIP_TYPE_RANGE:
            continue
        level = int(it["Level"])
        prices = {g: default_price(level, i) for i, g in enumerate(GRADES)}
        rows.append(
            {
                "recipe_id": r["rid"],
                "recipe_name": r["name"],
                "make_id": r["makeid"],
                "name": it["Secretname"],
                "type": itype,
                "type_name": EQUIP_TYPES.get(itype, str(itype)),
                "level": level,
                "skillid": r["skillid"],
                "typelv": r["typelv"],
                "prices": prices,
            }
        )

    rows.sort(key=lambda x: (x["type"], x["level"], x["make_id"]))

    wb = Workbook()
    default_ws = wb.active
    wb.remove(default_ws)

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(name="微软雅黑", bold=True, color="FFFFFF", size=11)
    thin = Side(style="thin", color="D0D7DE")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")

    # ---- 说明页 ----
    ws_info = wb.create_sheet("说明")
    ws_info.column_dimensions["A"].width = 22
    ws_info.column_dimensions["B"].width = 76
    title = ws_info.cell(row=1, column=1, value="可制造装备上架表单（默认定价）")
    title.font = Font(name="微软雅黑", bold=True, size=16, color="1F4E79")
    ws_info.merge_cells("A1:B1")
    ws_info.row_dimensions[1].height = 30
    notes = [
        ("来源", "制造配方表 other_tbrecipeinfodescconfig（skillid≤214 为装备）"),
        ("装备数", f"{len(rows)} 件可制造装备（全部可出售 Cansell=True）"),
        ("品级档", "D / C / B / A / S，对应客户端装备评分 equipscore1~5 图标（Id1=D最低 … Id5=S最高）"),
        ("品级阈值", f"{score_levels}  （评分百分比 < 阈值取该档，>100% 即 S 档）"),
        ("定价规则", "D 级 = 等级² × 1000（取整）；C/B/A/S 每上一档 ×1.10（对 D 逐档累乘后取整）"),
        ("默认货币", "金币（上架协议 Tcost=0，钻石 Tcost=1）"),
        ("编辑", "可在「装备表单」页直接改 D/C/B/A/S 列价格后另存，助手读取该 Excel 配置即可"),
        ("重复生成", "重新运行本脚本会覆盖本文件；用户改价后请勿再运行"),
    ]
    for i, (k, v) in enumerate(notes, start=3):
        c1 = ws_info.cell(row=i, column=1, value=k)
        c2 = ws_info.cell(row=i, column=2, value=v)
        c1.font = Font(name="微软雅黑", bold=True, size=10, color="1F4E79")
        c1.fill = PatternFill("solid", fgColor="E8F1F8")
        c1.border = border
        c1.alignment = center
        c2.font = Font(name="微软雅黑", size=10)
        c2.border = border
        c2.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws_info.row_dimensions[i].height = 32 if i <= 8 else 24

    # ---- 装备表单 ----
    columns = [
        ("type_name", "分类", 8),
        ("name", "装备名称", 22),
        ("level", "等级", 8),
        ("make_id", "道具ID", 10),
        ("recipe_id", "配方ID", 10),
    ]
    for g in GRADES:
        columns.append((f"price_{g}", f"{g}级默认价", 12))
    columns += [("currency", "默认货币", 10)]

    ws = wb.create_sheet("装备表单")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(columns))
    t = ws.cell(row=1, column=1, value="可制造装备 — 各品级默认上架价（金币）")
    t.font = Font(name="微软雅黑", bold=True, size=14, color="1F4E79")
    t.alignment = left
    ws.row_dimensions[1].height = 28
    meta = ws.cell(row=2, column=1, value=f"导出时间：{__import__('datetime').datetime.now():%Y-%m-%d %H:%M:%S}    条目数：{len(rows)}    D=等级²×1000，每品级+10%")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(columns))
    meta.font = Font(name="微软雅黑", size=9, color="666666")
    meta.alignment = left
    ws.row_dimensions[2].height = 20

    header_row = 3
    for col_idx, (_, title_cn, width) in enumerate(columns, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=title_cn)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = border
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[header_row].height = 22

    body_font = Font(name="微软雅黑", size=10)
    for row_idx, row in enumerate(rows, start=header_row + 1):
        fill = PatternFill("solid", fgColor="F2F7FB") if (row_idx - header_row) % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
        for col_idx, (key, _, _) in enumerate(columns, start=1):
            if key == "currency":
                value = "金币"
            elif key.startswith("price_"):
                value = row["prices"][key[6:]]
            else:
                value = row[key]
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = body_font
            cell.fill = fill
            cell.border = border
            if key in ("type_name", "name"):
                cell.alignment = left
            else:
                cell.alignment = center
                if isinstance(value, int):
                    cell.number_format = "0"

    last = header_row + len(rows)
    ws.auto_filter.ref = f"A{header_row}:{get_column_letter(len(columns))}{last}"
    ws.freeze_panes = "A4"

    wb.save(OUT_XLSX)
    print(f"共 {len(rows)} 件可制造装备 -> {OUT_XLSX}")
    print(f"品级阈值: {score_levels}")
    print("样例:")
    for row in rows[:5]:
        print(
            f"  [{row['type_name']}] {row['name']} Lv{row['level']} id={row['make_id']} "
            f"D={row['prices']['D']} C={row['prices']['C']} B={row['prices']['B']} "
            f"A={row['prices']['A']} S={row['prices']['S']}"
        )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
